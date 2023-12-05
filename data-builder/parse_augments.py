import openpyxl 
import json
import os

def parse_augments():
    wb = openpyxl.load_workbook(f"{os.path.dirname(__file__)}/augments.xlsx") 

    systems = {}
    
    for s in range(len(wb.sheetnames)):
        group = wb.sheetnames[s]

        wb.active = s
        ws = wb.active 

        system = group + ' Augment Slot'

        if system not in systems:
            systems[system] = {}
            systems[system]['*'] = []
            # We don't want the augments polluting the affix search
            systems[system]['hiddenFromAffixSearch'] = True

        rows = ws.iter_rows()
        next(rows) # skip header
        for row in rows:
            augmentName = row[1].value
            name = row[2].value
            level = row[0].value
            value = row[3].value
            bonusType = row[4].value
            description = row[5].value

            if level == 0:
                level = 1

            if augmentName is None and name is None and level is None and value is None:
                continue
        
            affix = {}
            affix['name'] = name
            affix['value'] = value
            affix['type'] = bonusType
            affix['description'] = description

            if affix['name'] in ['Fortification']:
                affix['value'] = 100 * affix['value']

            if len(systems[system]['*']) and 'name' in systems[system]['*'][-1] and augmentName == systems[system]['*'][-1]['name']:
                systems[system]['*'][-1]['affixes'].append(affix)
            else:
                option = {}
                option['affixes'] = [affix]
                option['ml'] = level
                if augmentName:
                    option['name'] = augmentName

                if 'name' in option and option['name'].startswith("Set Augment: "):
                    option['set'] = option['name'].replace("Set Augment: ", "")

                systems[system]['*'].append(option)
                


    return systems

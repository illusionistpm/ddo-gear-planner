from typing import cast

import openpyxl

import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from get_most_common_bonus_type import get_most_common_bonus_type
from write_json import write_json
import os
from parse_affixes_from_cell import canonicalize_affix_name


CANNITH_MAX_LEVEL = 34


def canonicalize_cannith_affix_name(affix: str) -> str:
    if affix.startswith('Insightful '):
        return 'Insightful ' + canonicalize_affix_name(affix.removeprefix('Insightful '))
    return canonicalize_affix_name(affix)

def parse_cannith() -> None:
    assumedBonusTypeMap = get_most_common_bonus_type()
    assumedBonusTypeMap['Perform'] = 'Enhancement'
    assumedBonusTypeMap['Songblade'] = 'Bool'

    wb = openpyxl.load_workbook(f"{os.path.dirname(__file__)}/cannith-crafting.xlsx")

    s = 0
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Sheet1':
            break
    wb.active = s

    ws = cast(openpyxl.worksheet.worksheet.Worksheet, wb.active)

    itemTypeInfoList = []

    levelStart = levelEnd = None
    words = []
    for idx, cell in enumerate(ws[1], 0):
        if isinstance(cell.value, str):
            words = cell.value.split()

        if cell.value == 1:
            levelStart = idx
        elif isinstance(cell.value, int):
            levelEnd = idx
        elif cell.value in ['Min Level', 'Sort']:
            continue
        elif words[-1] in ['Prefix', 'Suffix', 'Extra']:
            itemType = ' '.join(words[0:-1])
            itemTypeInfoList.append({'col': idx, 'itemType': itemType, 'affixLoc': words[-1]})
    assert levelStart is not None
    assert levelEnd is not None

    output = {}
    progression = {}
    itemTypes = {}
    output['progression'] = progression
    output['itemTypes'] = itemTypes
    output['maxLevel'] = CANNITH_MAX_LEVEL
    output['bonusTypes'] = assumedBonusTypeMap

    rows = ws.iter_rows()
    next(rows)
    for row in rows:
        affix = row[0].value
        assert isinstance(affix, str)
        affix = affix.replace('Ins.', 'Insightful')
        affix = affix.title()
        fixed_progression_value = None
        if affix == 'Songblade':
            fixed_progression_value = 1

        if affix == 'Spell Resistance (Sr)':
            affix = 'Spell Resistance'

        if affix == 'Resistance (Save)':
            affix = 'Resistance'

        if affix.startswith('Spell Focus: '):
            affix = affix.replace('Spell Focus: ', '') + ' Focus'

        affix = canonicalize_cannith_affix_name(affix)

        progVals = []
        for val in range(levelStart, levelEnd + 1):
            progVals.append(fixed_progression_value if fixed_progression_value is not None else row[val].value)

        progression[affix] = progVals

        for itemInfo in itemTypeInfoList:
            isMarked = row[itemInfo['col']].value
            if isMarked and len(isMarked) > 0:
                if itemInfo['itemType'] not in itemTypes:
                    itemTypes[itemInfo['itemType']] = {}

                if itemInfo['affixLoc'] not in itemTypes[itemInfo['itemType']]:
                    itemTypes[itemInfo['itemType']][itemInfo['affixLoc']] = []

                itemTypes[itemInfo['itemType']][itemInfo['affixLoc']].append(affix)

    # Only keep bonus types for things that are actually used by Cannith crafting
    delKeys = []
    for k,_ in assumedBonusTypeMap.items():
        if k not in progression:
            delKeys.append(k)

    for k in delKeys:
        del assumedBonusTypeMap[k]

    write_json(output, 'cannith')


if __name__ == "__main__":
    parse_cannith()

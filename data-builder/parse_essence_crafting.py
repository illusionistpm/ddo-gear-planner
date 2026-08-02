from typing import cast

from bs4 import BeautifulSoup
import openpyxl

import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from get_most_common_bonus_type import get_most_common_bonus_type
from write_json import write_json
import os
from parse_affixes_from_cell import canonicalize_affix_name


ESSENCE_CRAFTING_TABLE_PATH = f"{os.path.dirname(__file__)}/cache/crafting/Essence_Crafting_table_3b.html"


def canonicalize_essence_crafting_affix_name(affix: str) -> str:
    if affix.startswith('Insightful '):
        return 'Insightful ' + canonicalize_affix_name(affix.removeprefix('Insightful '))
    return canonicalize_affix_name(affix)


def normalize_wiki_stat_name(stat: str) -> str:
    stat = stat.replace('*', '').replace('Ins.', 'Insightful').strip()
    stat = stat.title()
    replacements = {
        'Alacrity Ranged/Melee': 'Alacrity',
        'Armor-Piercing': 'Armor-piercing',
        'Insightful Ench/Ill Resistance': 'Insightful Enchantment/Illusion Resistance',
        'Insightful Poi/Dis Ward': 'Insightful Poison/Disease Ward',
        'Insightful Spell Focus (One)': 'Insightful Spell Focus',
        'Insightful Spell Focus Mastery': 'Insightful Spell Focus Mastery',
        'Insightful Spellpower': 'Insightful Spell Power',
        'Insightful Vertigo /Stunning/Shatter': 'Insightful Vertigo/Stunning/Shatter',
        'Lore (All)': 'Universal Spell Lore',
        'Lore (One Type)': 'Lore',
        'Reflex/Fortitude/Will': 'Saves',
        'Resistance (Save)': 'Resistance',
        'Spell Focus (One Type)': 'Spell Focus',
        'Spell Focus Mastery': 'Spell Focus Mastery',
        'Spellpower': 'Spell Power',
        'Spellcasting Implement': 'Spellcasting Implement',
        'Spell Resistance (Sr)': 'Spell Resistance',
        'Weapon Dice Mult': 'Weapon Dice Multiplier',
    }
    return replacements.get(stat, stat)


def parse_wiki_value(value: str) -> int | float | str | None:
    value = value.strip()
    if value in ['', '-', '??', '?']:
        return None
    if 'd' in value:
        return value
    number = float(value)
    if number.is_integer():
        return int(number)
    return number


def get_essence_crafting_progression_from_wiki(soup: BeautifulSoup) -> tuple[list[int], dict[str, list[int | float | str | None]]]:
    data_table = None
    for table in soup.find_all('table'):
        first_row = table.find('tr')
        if not first_row:
            continue
        headers = [cell.get_text(' ', strip=True) for cell in first_row.find_all(['th', 'td'])]
        if headers and headers[0] == 'Min Level':
            data_table = table
            break

    if data_table is None:
        raise ValueError('Unable to find Essence Crafting table with Min Level header')

    rows = data_table.find_all('tr')
    level_cells = [cell.get_text(' ', strip=True) for cell in rows[0].find_all(['th', 'td'])][1:]
    levels = [int(level) for level in level_cells]

    progressions = {}
    for row in rows[1:]:
        cells = [cell.get_text(' ', strip=True) for cell in row.find_all(['th', 'td'])]
        if len(cells) != len(levels) + 1:
            continue
        progressions[normalize_wiki_stat_name(cells[0])] = [parse_wiki_value(value) for value in cells[1:]]

    return levels, progressions


def load_essence_crafting_progression_from_wiki(path: str = ESSENCE_CRAFTING_TABLE_PATH) -> tuple[list[int], dict[str, list[int | float | str | None]]]:
    with open(path, 'r', encoding='utf-8') as fh:
        soup = BeautifulSoup(fh.read(), 'html.parser')
    return get_essence_crafting_progression_from_wiki(soup)


def get_wiki_progression_key(affix: str) -> str | None:
    base_affix = affix.removeprefix('Insightful ')
    prefix = 'Insightful ' if affix.startswith('Insightful ') else ''

    if base_affix in ['Strength', 'Dexterity', 'Constitution', 'Intelligence', 'Wisdom', 'Charisma']:
        return f'{prefix}Ability'
    if base_affix in ['Enchantment Resistance', 'Illusion Resistance']:
        return f'{prefix}Enchantment/Illusion Resistance'
    if base_affix in ['Poison Ward', 'Disease Ward']:
        return f'{prefix}Poison/Disease Ward'
    if base_affix in ['Fortitude', 'Reflex', 'Will']:
        return 'Saves'
    if base_affix in ['Shatter', 'Stunning', 'Vertigo']:
        return f'{prefix}Vertigo/Stunning/Shatter'
    if base_affix in ['Acid Resistance', 'Cold Resistance', 'Electric Resistance', 'Fire Resistance',
                      'Light Resistance', 'Negative Resistance', 'Poison Resistance', 'Sonic Resistance']:
        return 'Resistance'
    if base_affix in ['Healing Amplification', 'Negative Amplification', 'Repair Amplification']:
        return 'Amplification'
    if base_affix in ['Acid Lore', 'Cold Lore', 'Fire Lore', 'Healing Lore', 'Ice Lore', 'Kinetic Lore',
                      'Lightning Lore', 'Negative Lore', 'Radiance Lore', 'Repair Lore', 'Sonic Lore', 'Void Lore']:
        return 'Lore'
    if base_affix in ['Acid Absorption', 'Cold Absorption', 'Electricity Absorption', 'Fire Absorption', 'Sonic Absorption']:
        return 'Absorption'
    if base_affix in ['Acid Spell Power', 'Fire Spell Power', 'Positive Spell Power', 'Repair Spell Power',
                      'Cold Spell Power', 'Kinetic Spell Power', 'Lightning Spell Power', 'Electric Spell Power',
                      'Negative Spell Power', 'Light Spell Power', 'Sonic Spell Power', 'Force Spell Power',
                      'Radiance', 'Reconstruction']:
        return f'{prefix}Spell Power'
    if base_affix in ['Abjuration Focus', 'Conjuration Focus', 'Enchantment Focus', 'Evocation Focus',
                      'Illusion Focus', 'Necromancy Focus', 'Transmutation Focus']:
        return f'{prefix}Spell Focus'
    if base_affix in ['Balance', 'Bluff', 'Concentration', 'Diplomacy', 'Disable Device', 'Haggle',
                      'Heal', 'Hide', 'Intimidate', 'Jump', 'Listen', 'Move Silently', 'Open Lock',
                      'Perform', 'Repair', 'Search', 'Spellcraft', 'Spot', 'Swim', 'Tumble']:
        return f'{prefix}Skill'
    if affix in ['Melee Alacrity', 'Ranged Alacrity']:
        return 'Alacrity'
    if affix == 'Armor-Piercing':
        return 'Armor-piercing'
    if affix in ['Insightful Glactiation', 'Insightful Cold Spell Power']:
        return 'Insightful Spell Power'
    if affix in ['Glactiation', 'Cold Spell Power']:
        return 'Spell Power'
    if affix == 'Magical Sheltering' or affix == 'Physical Sheltering':
        return 'Sheltering'
    if affix == 'Insightful Magical Sheltering' or affix == 'Insightful Physical Sheltering':
        return 'Insightful Sheltering'
    if affix == 'Spell Penetration':
        return 'Penetration'
    if affix == 'Insightful Spell Penetration':
        return 'Insightful Penetration'
    return affix


def get_max_known_level(
    levels: list[int],
    wiki_keys_used: set[str],
    wiki_progression: dict[str, list[int | float | str | None]],
    minimum_known_level: int,
) -> int:
    max_known_level = minimum_known_level
    for index, level in enumerate(levels):
        if level <= minimum_known_level:
            continue
        if all(wiki_progression[key][index] is not None for key in wiki_keys_used):
            max_known_level = level
        else:
            break
    return max_known_level


def parse_essence_crafting() -> None:
    assumedBonusTypeMap = get_most_common_bonus_type()
    assumedBonusTypeMap['Perform'] = 'Enhancement'
    assumedBonusTypeMap['Songblade'] = 'Bool'
    wiki_levels, wiki_progression = load_essence_crafting_progression_from_wiki()

    wb = openpyxl.load_workbook(f"{os.path.dirname(__file__)}/essence-crafting.xlsx")

    s = 0
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Sheet1':
            break
    wb.active = s

    ws = cast(openpyxl.worksheet.worksheet.Worksheet, wb.active)

    itemTypeInfoList = []

    levelStart = levelEnd = None
    words = []
    for idx, cell in enumerate(ws[1], 0):
        if isinstance(cell.value, str):
            words = cell.value.split()

        if cell.value == 1:
            levelStart = idx
        elif isinstance(cell.value, int):
            levelEnd = idx
        elif cell.value in ['Min Level', 'Sort']:
            continue
        elif words[-1] in ['Prefix', 'Suffix', 'Extra']:
            itemType = ' '.join(words[0:-1])
            itemTypeInfoList.append({'col': idx, 'itemType': itemType, 'affixLoc': words[-1]})
    assert levelStart is not None
    assert levelEnd is not None

    output = {}
    progression = {}
    itemTypes = {}
    output['progression'] = progression
    output['itemTypes'] = itemTypes
    output['bonusTypes'] = assumedBonusTypeMap
    wiki_keys_used = set()

    rows = ws.iter_rows()
    next(rows)
    for row in rows:
        affix = row[0].value
        assert isinstance(affix, str)
        affix = affix.replace('Ins.', 'Insightful')
        affix = affix.title()
        fixed_progression_value = None
        if affix == 'Songblade':
            fixed_progression_value = 1

        if affix == 'Spell Resistance (Sr)':
            affix = 'Spell Resistance'

        if affix == 'Resistance (Save)':
            affix = 'Resistance'

        if affix.startswith('Insightful Spell Focus: '):
            affix = 'Insightful ' + affix.removeprefix('Insightful Spell Focus: ') + ' Focus'

        if affix.startswith('Spell Focus: '):
            affix = affix.replace('Spell Focus: ', '') + ' Focus'

        affix = canonicalize_essence_crafting_affix_name(affix)

        sheet_prog_vals = []
        for val in range(levelStart, levelEnd + 1):
            sheet_value = row[val].value
            sheet_prog_vals.append(fixed_progression_value if fixed_progression_value is not None else sheet_value)

        wiki_key = get_wiki_progression_key(affix)
        if wiki_key in wiki_progression and any(value is not None for value in wiki_progression[wiki_key]):
            wiki_keys_used.add(wiki_key)
            progVals = [
                wiki_value if wiki_value is not None else sheet_prog_vals[index] if index < len(sheet_prog_vals) else None
                for index, wiki_value in enumerate(wiki_progression[wiki_key])
            ]
        else:
            progVals = sheet_prog_vals

        progression[affix] = progVals

        for itemInfo in itemTypeInfoList:
            isMarked = row[itemInfo['col']].value
            if isMarked and len(isMarked) > 0:
                if itemInfo['itemType'] not in itemTypes:
                    itemTypes[itemInfo['itemType']] = {}

                if itemInfo['affixLoc'] not in itemTypes[itemInfo['itemType']]:
                    itemTypes[itemInfo['itemType']][itemInfo['affixLoc']] = []

                itemTypes[itemInfo['itemType']][itemInfo['affixLoc']].append(affix)

    # Only keep bonus types for things that are actually used by Essence Crafting.
    delKeys = []
    for k,_ in assumedBonusTypeMap.items():
        if k not in progression:
            delKeys.append(k)

    for k in delKeys:
        del assumedBonusTypeMap[k]

    max_level = get_max_known_level(wiki_levels, wiki_keys_used, wiki_progression, levelEnd - levelStart + 1)
    output['maxLevel'] = max_level
    for affix, values in progression.items():
        if len(values) < max_level:
            last_value = values[-1]
            values.extend([last_value] * (max_level - len(values)))
        progression[affix] = values[:max_level]

    write_json(output, 'essence-crafting')


if __name__ == "__main__":
    parse_essence_crafting()

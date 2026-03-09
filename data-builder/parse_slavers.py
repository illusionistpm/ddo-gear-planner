from typing import TypedDict, cast

import openpyxl
import os
from get_inverted_synonym_map import get_inverted_synonym_map

from typedefs import Affix, AffixesDict, SetDict

SystemDict = TypedDict('SystemDict', { '*': list[AffixesDict|SetDict] })

def parse_slavers_sets():
    wb = openpyxl.load_workbook(f"{os.path.dirname(__file__)}/slavers.xlsx") 
    
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == 'Set':
            wb.active = s
            ws = wb.active 

    sets = {}

    synMap = get_inverted_synonym_map()

    rows = ws.iter_rows()
    next(rows) # skip header
    for row in rows:
        name = cast(str, row[0].value)
        level = row[1].value
        threshold = row[2].value
        affix = row[3].value
        value = row[4].value
        bonusType = row[5].value

        if level == 28:
            name = 'Legendary ' + name

        if name not in sets:
            sets[name] = []

        if len(sets[name]) == 0 or sets[name][-1]['threshold'] != threshold:
            thr = {}
            thr['threshold'] = threshold
            sets[name].append(thr)
        
        if not 'affixes' in sets[name][-1]:
            sets[name][-1]['affixes'] = []

        if affix in synMap:
            affix = synMap[affix]

        aff = {}
        aff['name'] = affix
        aff['value'] = value
        aff['type'] = bonusType

        sets[name][-1]['affixes'].append(aff)

    return sets


def parse_slavers_crafting() -> dict[str, SystemDict]:
    wb = openpyxl.load_workbook(f"{os.path.dirname(__file__)}/slavers.xlsx") 

    systems: dict[str, SystemDict] = {}
    
    for s in range(len(wb.sheetnames)):
        group = wb.sheetnames[s]

        wb.active = s
        ws = wb.active
        assert ws is not None

        if group == 'Set':
            continue

        system = "Slaver's " + group + ' Slot'
        legendarySystem = 'Legendary ' + system

        if system not in systems:
            systems[system] = { '*': [] }

        if legendarySystem not in systems:
            systems[legendarySystem] = { '*': [] }

        rows = ws.iter_rows()
        next(rows) # skip header
        for row in rows:
            name = cast(str, row[0].value)
            level = row[1].value
            value = row[2].value
            bonusType = cast(str, row[3].value)
        
            affix: Affix = {
                'name': name,
                'value': value,
                'type': bonusType,
            }

            if affix['name'] == 'Fortification':
                affix['value'] = 100 * affix['value']

            if affix['name'] == 'Use Magic Device (UMD)':
                affix['name'] = 'Use Magic Device'

            if affix['name'] == 'PRR':
                affix['name'] = 'Physical Resistance Rating'

            if affix['name'] == 'MRR':
                affix['name'] = 'Magical Resistance Rating'

            mySystem = system if level == 8 else legendarySystem
            
            systems[mySystem]['*'].append({ 'affixes': [affix] })

    systems["Slaver's Set Bonus"] = { '*': [] }
    systems["Legendary Slaver's Set Bonus"] =  { '*': [] }

    for setName in ["Slave Lord's Might", "Slave Lord's Sorcery", "Slave Lord's Endurance"]:
        systems["Slaver's Set Bonus"]['*'].append({ 'set': setName })
        systems["Legendary Slaver's Set Bonus"]['*'].append({ 'set': 'Legendary ' + setName })

    return systems

